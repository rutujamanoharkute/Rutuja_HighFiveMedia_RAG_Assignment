from azure.storage.blob import BlobServiceClient, ContainerClient
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader, TextLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.llms import Ollama
from langchain.chains import LLMChain
from langchain.prompts import PromptTemplate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime, timedelta
import tempfile
import os
import re
from collections import defaultdict
import hashlib
import io
import asyncio
from typing import List, Dict
import time

# Batch processing configuration
BATCH_SIZE = 4  # Number of documents to process in each batch
BATCH_TIMEOUT = 0.5  # Max seconds to wait before processing a partial batch

def get_ollama_llm():
    try:
        print("[INFO] Connecting to primary Ollama model...")
        OLLAMA_HOST = os.getenv('OLLAMA_HOST')
        MODEL_NAME = os.getenv('OLLAMA_MODEL')

        llm = Ollama(
            model=MODEL_NAME,
            base_url=OLLAMA_HOST,
            timeout=600,  # Increased timeout for batch processing
            temperature=0.1
        )
        llm.invoke("ping")
        print("[SUCCESS] Connected to primary Ollama model.")
        return llm

    except Exception as e:
        print(f"[WARN] Failed to connect to primary Ollama model: {e}")
        try:
            OLLAMA_HOST = os.getenv('OLLAMA_HOST_LOCAL')
            MODEL_NAME = os.getenv('OLLAMA_MODEL')

            llm = Ollama(
                model=MODEL_NAME,
                base_url=OLLAMA_HOST,
                timeout=600,
                temperature=0.1
            )
            llm.invoke("ping")
            return llm

        except Exception as e:
            print(f"[ERROR] Failed to connect to fallback Ollama model: {e}")
            raise RuntimeError("Both primary and fallback Ollama models are unreachable.")

# Initialize LLM
llm = get_ollama_llm()
# Improved analysis prompt with clear structure for parsing
analysis_prompt = """
You are a document analysis agent specialized in policy review. Your task is to review organizational policy documents and identify:

1. Document metadata:
   - Extract the exact policy title
   - Extract any dates including created date, last updated date, and expiration date (if available)
   - Extract document version number (if available)

2. Policy content analysis:
   - Identify outdated policies (any policy with expiration date in the past or content that mentions outdated regulations)
   - Extract the key policy topics and main sections

Document Content:
{document}

Respond in this exact format with appropriate values (leave empty if not found):

TITLE: [Full policy title]
CREATED_DATE: [Date in YYYY-MM-DD format]
UPDATED_DATE: [Date in YYYY-MM-DD format]
EXPIRATION_DATE: [Date in YYYY-MM-DD format]
VERSION: [Version number]
KEY_TOPICS: [Comma-separated list of main topics covered]
MAIN_SECTIONS: [Comma-separated list of main section titles]
OUTDATED_ELEMENTS: [Description of any outdated elements found]
POLICY_SUMMARY: [Brief 2-3 sentence summary of what this policy covers]
"""


# Batch processor class
class DocumentBatchProcessor:
    def __init__(self):
        self.current_batch = []
        self.last_batch_time = time.time()
        self.prompt_template = PromptTemplate(
            template=analysis_prompt,
            input_variables=["document"]
        )
    
    def add_document(self, content: str, metadata: dict) -> dict:
        """Add document to current batch and process if batch is ready"""
        self.current_batch.append((content, metadata))
        
        # Process batch if full or timeout reached
        if (len(self.current_batch) >= BATCH_SIZE or 
            (time.time() - self.last_batch_time) >= BATCH_TIMEOUT):
            return self.process_batch()
        return None
    
    def process_batch(self) -> List[Dict]:
        """Process the current batch of documents"""
        if not self.current_batch:
            return []
            
        print(f"[BATCH] Processing {len(self.current_batch)} documents")
        
        batch_contents = [item[0] for item in self.current_batch]
        batch_metadata = [item[1] for item in self.current_batch]
        
        try:
            # Prepare and run batch prompts
            chain = LLMChain(llm=llm, prompt=self.prompt_template)
            responses = chain.apply([{"document": content[:8000]} for content in batch_contents])
            
            # Parse responses and combine with metadata
            results = []
            for response, metadata in zip(responses, batch_metadata):
                doc_data = extract_structured_data(response['text'])
                doc_data.update(metadata)
                results.append(doc_data)
            
            return results
            
        except Exception as e:
            print(f"[ERROR] Batch processing failed: {e}")
            raise
        finally:
            self.current_batch = []
            self.last_batch_time = time.time()



blob_service_client = BlobServiceClient(
    account_url=f"https://{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}.blob.core.windows.net",
    credential=os.getenv('AZURE_STORAGE_ACCOUNT_KEY')
)

def download_from_blob(blob_name: str) -> str:
    """Download file from blob storage to temp location"""
    try:
        uploads_client = blob_service_client.get_container_client("uploads")
        blob_client = uploads_client.get_blob_client(blob_name)
        
        ext = os.path.splitext(blob_name)[1]
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as temp_file:
            download_stream = blob_client.download_blob()
            temp_file.write(download_stream.readall())
            print(f"[SUCCESS] Blob downloaded: {temp_file.name}")
            return temp_file.name
    except Exception as e:
        print(f"[ERROR] Failed to download blob {blob_name}: {str(e)}")
        raise RuntimeError(f"Failed to download blob {blob_name}: {str(e)}")

def parse_date(date_str):
    """Parse date string from various formats"""
    if not date_str or date_str.strip() == "" or date_str == "None" or date_str == "N/A":
        return None
    
    # Try common date formats
    date_formats = [
        "%Y-%m-%d", 
        "%m/%d/%Y", 
        "%d/%m/%Y", 
        "%B %d, %Y", 
        "%d %B %Y",
        "%Y/%m/%d"
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    print(f"[WARN] Unable to parse date: {date_str}")
    # If date parsing fails, return None
    return None

def calculate_content_hash(text):
    """Calculate a hash based on document content to identify similar documents"""
    return hashlib.md5(text.lower().encode('utf-8')).hexdigest()

def extract_structured_data(response_text):
    """Extract structured data from LLM response"""
    data = {}
    
    # Define patterns to extract fields
    patterns = {
        "title": r"TITLE:\s*(.*?)(?:\n|$)",
        "created_date": r"CREATED_DATE:\s*(.*?)(?:\n|$)",
        "updated_date": r"UPDATED_DATE:\s*(.*?)(?:\n|$)",
        "expiration_date": r"EXPIRATION_DATE:\s*(.*?)(?:\n|$)",
        "version": r"VERSION:\s*(.*?)(?:\n|$)",
        "key_topics": r"KEY_TOPICS:\s*(.*?)(?:\n|$)",
        "main_sections": r"MAIN_SECTIONS:\s*(.*?)(?:\n|$)",
        "outdated_elements": r"OUTDATED_ELEMENTS:\s*(.*?)(?:\n|$)",
        "policy_summary": r"POLICY_SUMMARY:\s*(.*?)(?:\n|$)"
    }
    
    # Extract each field
    for key, pattern in patterns.items():
        match = re.search(pattern, response_text, re.IGNORECASE | re.DOTALL)
        if match:
            data[key] = match.group(1).strip()
        else:
            data[key] = ""
    
    return data

def analyze_documents_from_blob(container_client, organizational_values):
    """Analyze documents from blob storage with batching support"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Analysis"
    
    # Set up headers with formatting (same as before)
    headers = ["File", "Title", "Created Date", "Updated Date", "Expiration Date", 
               "Version", "Status", "Issue Type", "Description", "Suggested Action", "Priority"]
    
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    documents_data = []
    file_content_map = {}
    content_hashes = defaultdict(list)
    today = datetime.now().date()
    row = 2

    # Initialize batch processor
    batch_processor = DocumentBatchProcessor()

    # Process each blob in the container
    for blob in container_client.list_blobs():
        try:
            print(f"[INFO] Processing document: {blob.name}")
            temp_path = download_from_blob(blob.name)
            ext = os.path.splitext(blob.name)[1].lower()
            
            loader_class = {
                '.pdf': PyPDFLoader,
                '.docx': Docx2txtLoader,
                '.txt': TextLoader
            }.get(ext)
            
            if not loader_class:
                print(f"[WARN] Unsupported file type {ext} for {blob.name}")
                continue
                
            loader = loader_class(temp_path)
            docs = loader.load()
            full_content = "\n".join([doc.page_content for doc in docs])
            content_hash = calculate_content_hash(full_content)
            content_hashes[content_hash].append(blob.name)
            file_content_map[blob.name] = full_content
            
            # Add document to batch processor
            metadata = {
                'filename': blob.name,
                'content_hash': content_hash
            }
            batch_result = batch_processor.add_document(full_content, metadata)
            
            # If batch was processed, add results
            if batch_result:
                documents_data.extend(batch_result)
            
            os.unlink(temp_path)
            
        except Exception as e:
            print(f"[ERROR] Error processing {blob.name}: {str(e)}")
            continue
    
    # Process any remaining documents in the batch
    final_batch = batch_processor.process_batch()
    if final_batch:
        documents_data.extend(final_batch)
    
    # Process each document and add issues to spreadsheet (same as before)
    for doc_data in documents_data:
        filename = doc_data['filename']
        display_filename = os.path.basename(filename)
        
        title = doc_data.get('title', '')
        created_date = doc_data.get('created_date', '')
        updated_date = doc_data.get('updated_date', '')
        expiration_date = doc_data.get('expiration_date', '')
        version = doc_data.get('version', '')
        
        parsed_expiration = parse_date(expiration_date)
        status = "Current"
        if parsed_expiration and parsed_expiration < today:
            status = "Expired"
        
        ws.append([
            display_filename,
            title,
            created_date,
            updated_date,
            expiration_date,
            version,
            status,
            "", "", "", ""
        ])
        
        if parsed_expiration and parsed_expiration < today:
            row += 1
            ws.append([
                display_filename, title, "", "", "", "", "",
                "Outdated Policy",
                f"Policy has expired on {expiration_date}",
                "Review and update policy with new expiration date",
                "High"
            ])
        
        if doc_data.get('outdated_elements', '').strip():
            row += 1
            ws.append([
                display_filename, title, "", "", "", "", "",
                "Outdated Content",
                doc_data['outdated_elements'],
                "Update policy content to meet current standards",
                "Medium"
            ])
        
        content_hash = doc_data['content_hash']
        if len(content_hashes[content_hash]) > 1:
            duplicate_files = [os.path.basename(f) for f in content_hashes[content_hash] if f != filename]
            if duplicate_files:
                row += 1
                ws.append([
                    display_filename, title, "", "", "", "", "",
                    "Redundant Policy",
                    f"This policy is identical or very similar to: {', '.join(duplicate_files)}",
                    "Consider consolidating redundant policies",
                    "Medium"
                ])
        
        row += 1
    
    # Set column widths (same as before)
    column_widths = {
        1: 40, 2: 30, 3: 15, 4: 15, 5: 15, 
        6: 10, 7: 12, 8: 20, 9: 50, 10: 50, 11: 12
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"policy_analysis_report_{timestamp}.xlsx"
    print(f"[SUCCESS] Report saved to {filename}")
    return excel_buffer, filename
